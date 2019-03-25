from threading import Thread
from openpyxl import load_workbook
import os,re
import sys
import subprocess
import threading
wb = load_workbook(sys.argv[3])
ports=[]
def get_col():
    print(wb.get_sheet_names())
    sheet_obj = wb.get_sheet_by_name(sys.argv[4])
    max_col=sheet_obj.max_column
    max_row=sheet_obj.max_row

    prot_col=""
    port_col=""
    for i in range(1, max_col + 1): 
        cell_obj = sheet_obj.cell(row = 1, column = i) 
        if cell_obj.value=='Protocol':
            prot_col=i
        if cell_obj.value=='LA Local Port':
            port_col=i

#def get_prot_port():
    for i in range(2,max_row+1):
        port_cell_obj = sheet_obj.cell(row=i,column=port_col)
        prot_cell_obj = sheet_obj.cell(row=i,column=prot_col)

        prot_str = prot_cell_obj.value
    
        if (prot_str.find("TCP") == 0):
            ports.append(int(port_cell_obj.value))
get_col()    
for k in ports:
	print (k)

def func1():
	
	for k in ports:	
		cmd = "dd if=/dev/urandom | ncat -v  "
		print(cmd+sys.argv[1]+" "+str(k))
		for j in range(1,int(sys.argv[5])):
			try:
				t1=subprocess.check_output(cmd+sys.argv[1]+" "+str(k),shell=True, stderr=subprocess.PIPE)
				print(t1)
			except subprocess.CalledProcessError,e:
				print(e.output)

        
#Thread(target=func1).start()



func1()



#import os
#import subprocess
#import sys
#import time

#pod = sys.argv[2]
#def compare():
#	f = open("Verdict.txt","w+")

#	cmd1 = "kubectl top pods "
#	cmd2 = "kubectl get pods | grep "
#
#	cpu_cmd = " | awk NR==2'{print $2+0}'"
#	mem_cmd = " | awk NR==2'{print $3+0}'"
#	status = " | awk '{print $3}'"
#	restart = " | awk '{print $4}'"
#	
#	cpumem = " | awk NR==2'{print $2+0,$3+0}' OFS='\t' >> cpumem.txt"
#	state_re = " | awk '{print $3,$4}' OFS='\t' >> restart.txt"


#	i=1
#	try:
#		while True:
 #   			cpu_mem = subprocess.check_output(cmd1+pod+cpumem, shell=True, stderr=subprocess.PIPE)
  #  			state = subprocess.check_output(cmd2+pod+state_re, shell=True, stderr=subprocess.PIPE)
   # 			s = subprocess.check_output(cmd2+pod+status, shell=True, stderr=subprocess.PIPE)
    #			r = int(float(subprocess.check_output(cmd2+pod+restart, shell=True, stderr=subprocess.PIPE)))
    #			c = int(float(subprocess.check_output(cmd1+pod+cpu_cmd, shell=True, stderr=subprocess.PIPE)))
    #			m = int(float(subprocess.check_output(cmd1+pod+mem_cmd, shell=True, stderr=subprocess.PIPE)))
    #			if ( c >= 50 or m >= 50 or r!=0 or s.strip()!="Running"):
     #   			f.write("FAIL\n")
      #  			break
    #			else:
     #   			print("Pass")
   #			time.sleep(5)
   #			i = i + 1
    #			if(i>=10):
     #   			break
#
#		f.write("PASS\n")
#	except ValueError,e:
#		print(e)
#


#if __name__ == "__main__":
 #   # creating thread
  #  t1 = threading.Thread(target=func1)
   # t2 = threading.Thread(target=compare)
#
    # starting thread 1
 #   t1.start()
    # starting thread 2
  #  t2.start()

    # wait until thread 1 is completely executed
   # t1.join()
    # wait until thread 2 is completely executed
   # t2.join()

    # both threads completely executed
   # print("Done!")

